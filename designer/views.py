# for excel:
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill

import json

from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.cache import never_cache
from django.conf import settings
import csv
from .services import run_analysis, UserInputError, normalize_input_id  # for errors
from .result_store import (
    save_analysis_results,
    get_full_rows,
    get_filtered_rows,
    get_form_data,
    load_analysis_results,
)
from .pipeline import apply_duplicate_mode, normalize_duplicate_mode
from .coverage_stats import build_coverage_figure, compute_coverage_stats
from .protein_map import (
    alphafold_pdb_path,
    alphafold_pdb_url,
    resolve_alphafold_pdb_url,
    resolve_domain_layout,
    resolve_plot_gene_label,
)
from .screen_plot import (
    SCREEN_MARKERS,
    build_overview_figure,
    get_screen_data_store,
    lookup_gene_by_uniprot,
    normalize_uniprot_accession,
    screen_plot_eligibility,
)
from .screen_plot.figure_cache import (
    cache_key,
    get_build_error,
    get_cached_figure,
    is_building,
    start_full_figure_build,
)
import math
from pathlib import Path

DEFAULT_COLUMNS = [
    'position', 'wt_codon', 'wt_aa', 'mut_aa',
    'outcomes', 'avg_alpha_score', 'alpha_score',
    'sgrna_seq', 'protospacer',
    'pam', 'strand', 'target_position',
]

COLUMN_LABELS = {
    'position': 'Position',
    'wt_codon': 'WT Codon',
    'wt_aa': 'WT AA',
    'mut_aa': 'Mutated AA',
    'outcomes': 'Guide RNA Outcomes',
    'avg_alpha_score': 'avg. AlphaMissense Score',
    'alpha_score': 'AlphaMissense Score',
    'editor_used': 'Editor',
    'sgrna_seq': 'Guide RNA',
    'protospacer': 'Protospacer',
    'pam': 'PAM',
    'strand': 'Strand',
    'target_position': 'Target Position',
}

DISPLAY_COLUMNS_MAP = [
    ('position', 'Position'),
    ('wt_codon', 'WT Codon'),
    ('wt_aa', 'WT AA'),
    ('mut_aa', 'Mutated AA'),
    ('outcomes', 'Outcomes'),
    ('avg_alpha_score', 'avg. AlphaMissense Score'),
    ('alpha_score', 'AlphaMissense Score'),
    ('editor_used', 'Editor'),
    ('sgrna_seq', 'Guide RNA'),
    ('protospacer', 'Protospacer'),
    ('pam', 'PAM'),
    ('strand', 'Strand'),
    ('target_position', 'Target Pos.'),
]

EXPORT_COLUMN_MAPPING = {
    'position': ('Position', 'position'),
    'wt_codon': ('WT Codon', 'wt_codon'),
    'wt_aa': ('WT AA', 'wt_aa'),
    'mut_aa': ('Mutated AA', 'mut_aa'),
    'alpha_score': ('AlphaMissense Score', 'alpha_score'),
    'editor_used': ('Editor', 'editor_used'),
    'sgrna_seq': ('sgRNA Sequence', 'sgrna_seq'),
    'protospacer': ('Protospacer', 'protospacer'),
    'pam': ('PAM', 'pam'),
    'strand': ('Strand', 'strand'),
    'target_position': ('Target Position', 'target_position'),
    'outcomes': ('Guide RNA Outcomes', 'outcomes'),
    'avg_alpha_score': ('avg. AlphaMissense Score', 'avg_alpha_score'),
}


def parse_editing_window(post):
    window_min_raw = post.get('window_min')
    window_max_raw = post.get('window_max')
    try:
        window_min = int(window_min_raw) if window_min_raw not in [None, '', 'None'] else None
        window_max = int(window_max_raw) if window_max_raw not in [None, '', 'None'] else None
    except (TypeError, ValueError):
        window_min = None
        window_max = None
    return window_min, window_max


def resolve_score_filter(post, default_cutoff='0.5'):
    """Return (score_filter_mode, display_score_cutoff) for results display."""
    mode = post.get('score_filter_mode')
    cutoff_raw = post.get('display_score_cutoff')

    if mode is None:
        legacy = post.get('filter_threshold')
        if legacy == 'true':
            mode = 'above'
        elif legacy == 'false':
            mode = 'all'
        else:
            output_mode = post.get('output')
            if output_mode == 'only_over_threshold':
                mode = 'above'
            elif output_mode == 'all':
                mode = 'all'
            else:
                mode = 'above'

    if mode not in ('all', 'above', 'below'):
        mode = 'all'

    if cutoff_raw in [None, '']:
        cutoff_raw = post.get('alpha_threshold', default_cutoff)
    try:
        cutoff = float(cutoff_raw)
    except (TypeError, ValueError):
        cutoff = float(default_cutoff)

    return mode, cutoff


def resolve_selected_columns(post, editor):
    selected_columns = post.getlist('columns')
    if not selected_columns:
        selected_columns = list(DEFAULT_COLUMNS)
    if editor == 'BOTH' and 'editor_used' not in selected_columns:
        if 'alpha_score' in selected_columns:
            idx = selected_columns.index('alpha_score') + 1
            selected_columns.insert(idx, 'editor_used')
        else:
            selected_columns.append('editor_used')
    return selected_columns


def _row_has_valid_score(row):
    score = row.get('alpha_score')
    if score is None:
        return False
    if isinstance(score, float) and math.isnan(score):
        return False
    return True


def filter_rows_by_score(rows, score_filter_mode, display_score_cutoff):
    if score_filter_mode == 'all':
        return rows
    try:
        cutoff = float(display_score_cutoff)
    except (TypeError, ValueError):
        return rows

    if score_filter_mode == 'below':
        return [
            row for row in rows
            if _row_has_valid_score(row) and float(row['alpha_score']) <= cutoff
        ]

    # default: 'above'
    return [
        row for row in rows
        if _row_has_valid_score(row) and float(row['alpha_score']) >= cutoff
    ]


def build_form_data(**kwargs):
    return kwargs


def _normalize_analysis_result(analysis_result):
    if isinstance(analysis_result, dict):
        result = dict(analysis_result)
        result.setdefault('guide_rows', [])
        result.setdefault('no_guide_positions', [])
        return result
    return {'guide_rows': analysis_result, 'no_guide_positions': []}


def _coverage_context(result_rows, form_data, no_guide_rows=None, coverage_rows=None):
    """Summary stats + Plotly coverage figure for the results page.

    coverage_rows should be duplicate-mode-applied but score-unfiltered so the
    bar can show both ≥ and < cutoff fragments.
    """
    try:
        protein_length = int(form_data.get('protein_length') or 0)
    except (TypeError, ValueError):
        protein_length = 0
    try:
        cutoff = float(
            form_data.get(
                'display_score_cutoff',
                form_data.get('alpha_threshold', 0.5),
            )
        )
    except (TypeError, ValueError):
        cutoff = 0.5

    rows_for_bar = coverage_rows if coverage_rows is not None else result_rows
    stats = compute_coverage_stats(
        rows_for_bar,
        protein_length,
        cutoff,
        guide_count=len(result_rows),
        no_guide_count=len(no_guide_rows or []),
    )
    # Editable count for the summary line reflects the current table filter.
    display_positions = {
        int(row['position'])
        for row in result_rows
        if row.get('position') is not None and row.get('sgrna_seq')
    }
    stats['editable_positions_display'] = len(display_positions)
    stats['coverage_pct_display'] = (
        round(100.0 * len(display_positions) / protein_length, 1) if protein_length else 0.0
    )

    figure_json = None
    if protein_length > 0:
        try:
            figure_json = json.loads(build_coverage_figure(stats).to_json())
        except Exception:
            figure_json = None

    guide_positions = sorted(display_positions)
    return {
        'coverage_stats': stats,
        'coverage_figure': figure_json,
        'coverage_figure_json': json.dumps(figure_json) if figure_json else 'null',
        'guide_positions_json': json.dumps(guide_positions),
        'structure_accession': normalize_uniprot_accession(
            form_data.get('uniprot_accession') or form_data.get('uniprot_id') or ''
        ),
    }


def _no_guide_for_display(no_guide_positions, form_data, post=None):
    """Apply score filter to positions without guides (no duplicate-mode)."""
    post_data = post if post is not None else form_data
    score_filter_mode, display_score_cutoff = resolve_score_filter(
        post_data,
        default_cutoff=form_data.get('alpha_threshold', '0.5'),
    )
    return filter_rows_by_score(no_guide_positions, score_filter_mode, display_score_cutoff)


def _screen_plot_context(form_data):
    uniprot_id = (
        form_data.get('uniprot_accession')
        or form_data.get('uniprot_id')
        or ''
    ).strip()
    status = screen_plot_eligibility(uniprot_id)
    if status['available']:
        get_screen_data_store().start_background_load()
    return {
        'screen_plot_available': status['available'],
        'screen_plot_gene': status.get('gene') or resolve_plot_gene_label(form_data),
        'screen_plot_markers': list(SCREEN_MARKERS),
        'screen_plot_lookup_uniprot': status.get('lookup_uniprot') or '',
        'screen_plot_unavailable_reason': status.get('reason'),
        'protein_map_available': True,
    }


def _build_results_context(result_rows, form_data, no_guide_rows=None, full_rows=None):
    unique_positions = sorted({row['position'] for row in result_rows}) if result_rows else []
    no_guide_rows = no_guide_rows or []
    # Coverage bar uses all guides after duplicate mode (no score filter) so
    # both ≥ and < cutoff fragments appear.
    duplicate_mode = normalize_duplicate_mode(form_data.get('duplicate_mode', 'best'))
    coverage_rows = apply_duplicate_mode(full_rows or result_rows, duplicate_mode)
    return {
        'results': result_rows,
        'position_count': len(unique_positions),
        'guide_count': len(result_rows),
        'no_guide_positions': no_guide_rows,
        'no_guide_count': len(no_guide_rows),
        'show_editor_in_no_guide': form_data.get('editor') == 'BOTH',
        'display_columns': [COLUMN_LABELS.get(col, col) for col in form_data.get('selected_columns', [])],
        'display_columns_map': DISPLAY_COLUMNS_MAP,
        'form_data': form_data,
        **_screen_plot_context(form_data),
        **_coverage_context(
            result_rows,
            form_data,
            no_guide_rows,
            coverage_rows=coverage_rows,
        ),
    }


def _rows_for_display(full_rows, form_data, post=None):
    """Apply duplicate handling then score filter for the results table."""
    post_data = post if post is not None else form_data
    duplicate_mode = normalize_duplicate_mode(
        post_data.get('duplicate_mode', form_data.get('duplicate_mode', 'best'))
    )
    rows = apply_duplicate_mode(full_rows, duplicate_mode)
    score_filter_mode, display_score_cutoff = resolve_score_filter(
        post_data,
        default_cutoff=form_data.get('alpha_threshold', '0.5'),
    )
    return filter_rows_by_score(rows, score_filter_mode, display_score_cutoff)


def _render_results(request, result_rows, full_rows, form_data, no_guide_full=None, no_guide_display=None):
    no_guide_full = no_guide_full if no_guide_full is not None else []
    no_guide_display = no_guide_display if no_guide_display is not None else []
    save_analysis_results(
        request,
        full_rows=full_rows,
        filtered_rows=result_rows,
        no_guide_positions=no_guide_full,
        form_data=form_data,
    )
    return render(
        request,
        'designer/results.html',
        _build_results_context(
            result_rows,
            form_data,
            no_guide_display,
            full_rows=full_rows,
        ),
    )


def _refresh_results_from_cache(request):
    cached = load_analysis_results(request)
    full_rows = cached.get('full_rows') or cached.get('raw_rows')
    if full_rows is None:
        return redirect('home')

    form_data = dict(cached.get('form_data') or {})
    editor = request.POST.get('editor', form_data.get('editor', ''))
    alpha_threshold = request.POST.get('alpha_threshold', form_data.get('alpha_threshold', '0.5'))
    score_filter_mode, display_score_cutoff = resolve_score_filter(
        request.POST, default_cutoff=form_data.get('alpha_threshold', alpha_threshold)
    )
    selected_columns = resolve_selected_columns(request.POST, editor)
    editor_display = 'ABE & CBE' if editor == 'BOTH' else editor

    result_rows = _rows_for_display(full_rows, form_data, post=request.POST)
    no_guide_full = cached.get('no_guide_positions', [])
    no_guide_display = _no_guide_for_display(no_guide_full, form_data, post=request.POST)

    form_data.update({
        'editor': editor,
        'editor_display': editor_display,
        'alpha_threshold': alpha_threshold,
        'selected_columns': selected_columns,
        'score_filter_mode': score_filter_mode,
        'display_score_cutoff': display_score_cutoff,
        'filter_threshold': 'true' if score_filter_mode == 'above' else 'false',
    })

    return _render_results(
        request, result_rows, full_rows, form_data,
        no_guide_full=no_guide_full,
        no_guide_display=no_guide_display,
    )


def _rows_for_export(request):
    form_data = get_form_data(request)
    full_rows = get_full_rows(request)
    if not full_rows:
        return [], form_data
    duplicate_mode = normalize_duplicate_mode(form_data.get('duplicate_mode', 'best'))
    rows = apply_duplicate_mode(full_rows, duplicate_mode)
    rows = apply_threshold_filter(rows, form_data)
    return rows, form_data


def home(request):
    return render(request, 'designer/home.html')

def loading(request):
    return render(request, 'designer/loading.html')

@never_cache
def results(request):
    if request.method != 'POST':
        return redirect('home')

    if request.POST.get('refresh_only') == '1':
        return _refresh_results_from_cache(request)

    uniprot_id = normalize_input_id(request.POST.get('uniprot_id', ''))
    editor = request.POST.get('editor', '')
    alpha_threshold = request.POST.get('alpha_threshold', '0.5')
    top_sgrnas = request.POST.get('top_sgrnas', '')
    duplicate_mode = normalize_duplicate_mode(request.POST.get('duplicate_mode', 'best'))
    pam_type = request.POST.get('pam_type', 'NGG')

    score_filter_mode, display_score_cutoff = resolve_score_filter(
        request.POST, default_cutoff=alpha_threshold
    )
    window_min, window_max = parse_editing_window(request.POST)
    selected_columns = resolve_selected_columns(request.POST, editor)
    editor_display = 'ABE & CBE' if editor == 'BOTH' else editor

    try:
        analysis_result = _normalize_analysis_result(run_analysis(
            uniprot_id=uniprot_id,
            editor=editor,
            alpha_threshold=alpha_threshold,
            top_sgrnas=top_sgrnas,
            window_min=window_min,
            window_max=window_max,
            pam_type=pam_type,
        ))
        full_rows = analysis_result['guide_rows']
        no_guide_full = analysis_result['no_guide_positions']
        form_data = build_form_data(
            uniprot_id=uniprot_id,
            editor=editor,
            editor_display=editor_display,
            alpha_threshold=alpha_threshold,
            top_sgrnas=top_sgrnas,
            gene_name=analysis_result.get('gene_symbol') or uniprot_id,
            gene_symbol=analysis_result.get('gene_symbol') or '',
            uniprot_accession=analysis_result.get('uniprot_accession') or uniprot_id,
            protein_length=analysis_result.get('protein_length') or 0,
            transcript_id=analysis_result.get('transcript_id') or '',
            selected_columns=selected_columns,
            score_filter_mode=score_filter_mode,
            display_score_cutoff=display_score_cutoff,
            filter_threshold='true' if score_filter_mode == 'above' else 'false',
            window_min=window_min,
            window_max=window_max,
            duplicate_mode=duplicate_mode,
            pam_type=pam_type,
        )
        result_rows = _rows_for_display(full_rows, form_data, post=request.POST)
        no_guide_display = _no_guide_for_display(no_guide_full, form_data, post=request.POST)

        return _render_results(
            request, result_rows, full_rows, form_data,
            no_guide_full=no_guide_full,
            no_guide_display=no_guide_display,
        )

    except UserInputError as e:
        return render(request, 'designer/error.html', {'error': str(e), 'error_type': 'user'})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return render(request, 'designer/error.html', {'error': str(e), 'error_type': 'server'})


def apply_threshold_filter(rows, form_data):
    mode = form_data.get('score_filter_mode')
    if mode is None:
        mode = 'above' if form_data.get('filter_threshold') == 'true' else 'all'
    cutoff = form_data.get(
        'display_score_cutoff',
        form_data.get('alpha_threshold', '0.5'),
    )
    return filter_rows_by_score(rows, mode, cutoff)

def download_excel(request):
    rows, form_data = _rows_for_export(request)
    duplicate_mode = normalize_duplicate_mode(form_data.get('duplicate_mode', 'best'))

    gene_name = form_data.get('gene_name') or form_data.get('uniprot_id', 'Unknown target')
    editor_display = form_data.get('editor_display', form_data.get('editor', ''))

    output_mode = form_data.get('output_mode', 'all')
    alpha_threshold = form_data.get('alpha_threshold')

    # Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'

    row_idx = 1


    # ===== titel =====
    title = f"CRISPR Base Editing Designer Results for {gene_name}"
    ws.cell(row=row_idx, column=1, value=title)
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 2

    # ===== Tabelle 1: Zusammenfassung =====
    headers_1 = ['UniProt ID', 'Editor', 'PAM', 'AlphaMissense Threshold', 'Top Guide RNAs']
    values_1 = [
        form_data.get('uniprot_id', ''),
        editor_display,
        form_data.get('pam_type', 'NGG'),
        form_data.get('alpha_threshold', ''),
        form_data.get('top_sgrnas', ''),
    ]

    for col, header in enumerate(headers_1, start=1):
        ws.cell(row=row_idx, column=col, value=header).font = Font(bold=True)

    row_idx += 1

    for col, value in enumerate(values_1, start=1):
        ws.cell(row=row_idx, column=col, value=value)

    row_idx += 3

    #tabel 2: results
    column_mapping = {
        'position': ('Position', 'position'),
        'wt_codon': ('WT Codon', 'wt_codon'),
        'wt_aa': ('WT AA', 'wt_aa'),
        'mut_aa': ('Mutated AA', 'mut_aa'),
        'alpha_score': ('AlphaMissense Score', 'alpha_score'),
        'editor_used': ('Editor', 'editor_used'),
        'sgrna_seq': ('sgRNA Sequence', 'sgrna_seq'),
        'protospacer': ('Protospacer', 'protospacer'),
        'pam': ('PAM', 'pam'),
        'strand': ('Strand', 'strand'),
        'target_position': ('Target Position', 'target_position'),
        'outcomes': ('Guide RNA Outcomes', 'outcomes'),
        'avg_alpha_score': ('avg. AlphaMissense Score','avg_alpha_score'),
    }

    selected_columns = form_data.get('selected_columns', [])

    # if nothing selected -> show all
    if not selected_columns:
        selected_columns = list(column_mapping.keys())

    # Header
    for col_idx, col_key in enumerate(selected_columns, start=1):
        header = column_mapping[col_key][0]
        ws.cell(row=row_idx, column=col_idx, value=header).font = Font(bold=True)

    row_idx += 1

    # colors for groups of occurrences of same sgRNA
    group_colors = [
        '86EFAC',  # soft green
        '93C5FD',  # soft blue
        'D8B4FE',  # soft purple
        'FDB474',  # soft orange
        'F9A8D4',  # soft pink
    ]

    # counts every sgRNA
    from collections import Counter
    seq_counts = Counter(r.get('sgrna_seq') for r in rows)

    # map color to same sgRNAs
    seq_color_map = {}
    color_index = 0
    for r in rows:
        seq = r.get('sgrna_seq')
        if seq and seq_counts[seq] > 1 and seq not in seq_color_map:
            seq_color_map[seq] = group_colors[color_index % len(group_colors)]
            color_index += 1

    for r in rows:
        seq = r.get('sgrna_seq')
        row_color = seq_color_map.get(seq) if duplicate_mode == 'group' else None

        for col_idx, col_key in enumerate(selected_columns, start=1):
            _, data_key = column_mapping[col_key]
            value = r.get(data_key, '')

            if data_key == 'target_position' and isinstance(value, list):
                value = ', '.join(str(x) for x in value)

            if data_key == 'outcomes' and isinstance(value, list):
                value = '\n'.join(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if data_key == 'outcomes':
                cell.alignment = Alignment(wrap_text=True)

            if row_color:
                cell.fill = PatternFill(
                    start_color=row_color,
                    end_color=row_color,
                    fill_type='solid'
                )

        row_idx += 1

    for i in range(1, len(selected_columns) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20

    # File:
    filename = f"{gene_name}_crispr_results.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


def download_csv(request):
    rows, form_data = _rows_for_export(request)

    output_mode = form_data.get('output_mode', 'all')
    alpha_threshold = form_data.get('alpha_threshold')


    column_mapping = {
        'position': ('Position', 'position'),
        'wt_codon': ('WT Codon', 'wt_codon'),
        'wt_aa': ('WT AA', 'wt_aa'),
        'mut_aa': ('Mutated AA', 'mut_aa'),
        'alpha_score': ('AlphaMissense Score', 'alpha_score'),
        'editor_used': ('Editor', 'editor_used'),
        'sgrna_seq': ('sgRNA Sequence', 'sgrna_seq'),
        'protospacer': ('Protospacer', 'protospacer'),
        'pam': ('PAM', 'pam'),
        'strand': ('Strand', 'strand'),
        'target_position': ('Target Position', 'target_position'),
        'outcomes': ('Guide RNA Outcomes', 'outcomes'),
        'avg_alpha_score': ('avg. AlphaMissense Score','avg_alpha_score'),
    }

    selected_columns = form_data.get('selected_columns', [])
    if not selected_columns:
        selected_columns = list(column_mapping.keys())


    gene_name = form_data.get('gene_name') or form_data.get('uniprot_id', 'Unknown target')

    filename = f"{gene_name}_crispr_results.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # header
    headers = [column_mapping[col][0] for col in selected_columns]
    writer.writerow(headers)

    # data
    for r in rows:
        csv_row = []
        for col_key in selected_columns:
            _, data_key = column_mapping[col_key]
            value = r.get(data_key, '')

            if data_key == 'target_position':
                if isinstance(value, list):
                    value = 'pos ' + ' / '.join(str(x) for x in value)
                elif value is None:
                    value = ''
                else:
                    value = f'pos {value}'

            if data_key == 'outcomes' and isinstance(value, list):
                value = ' | '.join(value)

            csv_row.append(value)

        writer.writerow(csv_row)

    return response


def _guide_positions_from_rows(rows) -> list[int]:
    """Unique amino-acid positions with a guide in the visible results table."""
    positions = set()
    for row in rows or []:
        if not row.get('sgrna_seq'):
            continue
        pos = row.get('position')
        if pos is None:
            continue
        try:
            positions.add(int(pos))
        except (TypeError, ValueError):
            continue
    return sorted(positions)


def screen_plot_status(request):
    """JSON readiness for background screen data load (avoids nginx timeout on first plot)."""
    store = get_screen_data_store()
    payload = {
        'ready': store.is_ready(),
        'loading': store.is_loading(),
        'error': store.get_load_error(),
        'stage': store.load_stage,
        'elapsed_sec': store.load_elapsed_sec(),
    }
    uniprot_id = (request.GET.get('uniprot_id') or '').strip()
    if uniprot_id:
        ctx, err = _full_screen_plot_context(request, uniprot_id)
        if err is not None:
            payload.update({
                'figure_ready': False,
                'figure_building': False,
                'figure_error': 'Could not resolve screen plot context.',
            })
        else:
            gene, guide_positions, _meta = ctx
            started = _ensure_full_figure_build_started(gene, guide_positions, store)
            key = cache_key(gene, guide_positions)
            payload.update({
                'figure_ready': get_cached_figure(key) is not None,
                'figure_building': is_building(key) or started,
                'figure_error': get_build_error(key),
            })
    return JsonResponse(payload)


@require_POST
def screen_plot_warmup(request):
    """Start background load; returns immediately."""
    store = get_screen_data_store()
    started = store.start_background_load()
    return JsonResponse({
        'started': started,
        'ready': store.is_ready(),
        'loading': store.is_loading(),
        'error': store.get_load_error(),
    })


@require_POST
def screen_plot_overview(request):
    """Domain + AMBER guide track below the results table (any analyzed protein)."""
    form_data = get_form_data(request)
    uniprot_id = normalize_input_id(
        request.POST.get('uniprot_id') or form_data.get('uniprot_id') or ''
    )
    accession = normalize_uniprot_accession(
        form_data.get('uniprot_accession') or uniprot_id
    )
    gene = resolve_plot_gene_label({
        **form_data,
        'uniprot_id': uniprot_id,
        'uniprot_accession': accession,
    })

    try:
        protein_length = int(form_data.get('protein_length') or 0)
    except (TypeError, ValueError):
        protein_length = 0

    if protein_length <= 0:
        meta = lookup_gene_by_uniprot(accession or uniprot_id)
        if meta:
            protein_length = int(meta.get('uniprot_len') or meta.get('protein_len') or 0)
            gene = meta.get('gene') or gene

    if protein_length <= 0:
        return JsonResponse(
            {
                'available': False,
                'reason': 'Protein length is not available for this result. Re-run the analysis.',
            },
            status=400,
        )

    domain_layout, domain_source = resolve_domain_layout(
        gene=gene,
        uniprot_accession=accession,
    )
    guide_positions = _guide_positions_from_rows(get_filtered_rows(request))
    store = get_screen_data_store()

    # Start screen data only when this gene is in the library (for guide coloring).
    status = screen_plot_eligibility(accession or uniprot_id)
    if status['available']:
        store.start_background_load()

    try:
        fig = build_overview_figure(
            gene,
            guide_positions=guide_positions,
            store=store,
            protein_length=protein_length,
            domain_layout=domain_layout,
        )
        figure_json = json.loads(fig.to_json())
    except Exception as exc:
        return JsonResponse(
            {'available': False, 'reason': f'Could not build overview: {exc}'},
            status=500,
        )

    return JsonResponse({
        'available': True,
        'figure': figure_json,
        'gene': gene,
        'colored_guides': store.is_ready() and status['available'],
        'domain_source': domain_source,
    })


@require_GET
def structure_pdb(request, accession: str):
    """Serve AlphaFold PDB same-origin (local file or proxied from AFDB)."""
    import requests as http_requests

    safe = normalize_uniprot_accession(accession)
    if not safe or '/' in safe or '\\' in safe or '..' in safe:
        raise Http404('Invalid accession')

    local = alphafold_pdb_path(safe)
    if local is not None:
        root = local.resolve().parent
        expected_root = (
            Path(getattr(settings, 'SCREEN_DATA_DIR', settings.BASE_DIR / 'files-archive-dir'))
            / 'libraries'
            / 'AF_screen_proteins'
        ).resolve()
        if root != expected_root or not local.resolve().is_file():
            raise Http404('Structure not found')
        return FileResponse(local.open('rb'), content_type='chemical/x-pdb', filename=local.name)

    # Proxy AFDB so Molstar gets same-origin PDB bytes (redirects often fail in the viewer).
    remote_url = resolve_alphafold_pdb_url(safe)
    if not remote_url:
        raise Http404(f'Structure not available for {safe}')
    try:
        remote = http_requests.get(remote_url, timeout=60)
        remote.raise_for_status()
    except Exception as exc:
        raise Http404(f'Structure not available for {safe}') from exc

    return HttpResponse(
        remote.content,
        content_type='chemical/x-pdb',
        headers={'Content-Disposition': f'inline; filename="AF-{safe}-F1-model.pdb"'},
    )


def _full_screen_plot_context(request, uniprot_id: str):
    """Shared eligibility + gene + guide positions for full-screen plot endpoints."""
    status = screen_plot_eligibility(uniprot_id)
    if status['reason'] == 'data_unavailable':
        return None, JsonResponse(
            {
                'available': False,
                'reason': 'Screen plot data files are not available on the server.',
            },
            status=503,
        )
    meta = status.get('meta') or lookup_gene_by_uniprot(uniprot_id)
    if meta is None:
        return None, JsonResponse(
            {
                'available': False,
                'reason': 'Your UniProt ID is not in the published NGG screen library.',
            },
            status=404,
        )
    gene = meta['gene']
    guide_positions = _guide_positions_from_rows(get_filtered_rows(request))
    return (gene, guide_positions, meta), None


def _ensure_full_figure_build_started(gene: str, guide_positions: list[int], store) -> bool:
    """Start background full-figure build if data is ready and nothing is cached/in-flight."""
    if not store.is_ready():
        return False
    key = cache_key(gene, guide_positions)
    if get_cached_figure(key) is not None or is_building(key) or get_build_error(key):
        return False
    return start_full_figure_build(gene, guide_positions, store=store)


@require_POST
def screen_plot_prewarm_full(request):
    """Build full-screen figure in background (avoids gateway timeout)."""
    uniprot_id = (request.POST.get('uniprot_id') or '').strip()
    ctx, err = _full_screen_plot_context(request, uniprot_id)
    if err is not None:
        return err
    gene, guide_positions, _meta = ctx
    store = get_screen_data_store()
    if not store.is_ready():
        if not store.is_loading():
            store.start_background_load()
        return JsonResponse({'ready': False, 'building': False, 'loading_data': True}, status=202)

    key = cache_key(gene, guide_positions)
    started = _ensure_full_figure_build_started(gene, guide_positions, store)
    return JsonResponse({
        'ready': get_cached_figure(key) is not None,
        'building': is_building(key) or started,
        'loading_data': False,
        'error': get_build_error(key),
    })


def screen_plot_full_status(request):
    """Poll readiness of cached full-screen figure."""
    uniprot_id = (request.GET.get('uniprot_id') or request.POST.get('uniprot_id') or '').strip()
    ctx, err = _full_screen_plot_context(request, uniprot_id)
    if err is not None:
        return err
    gene, guide_positions, _meta = ctx
    store = get_screen_data_store()
    key = cache_key(gene, guide_positions)
    started = _ensure_full_figure_build_started(gene, guide_positions, store)
    return JsonResponse({
        'ready': get_cached_figure(key) is not None,
        'building': is_building(key) or started,
        'loading_data': not store.is_ready(),
        'error': get_build_error(key),
    })


@require_POST
def screen_enrichment_plot(request):
    """Return cached full-screen Plotly figure (all markers, LFC only)."""
    uniprot_id = (request.POST.get('uniprot_id') or '').strip()

    ctx, err = _full_screen_plot_context(request, uniprot_id)
    if err is not None:
        return err

    gene, guide_positions, _meta = ctx
    store = get_screen_data_store()
    if not store.is_ready():
        if not store.is_loading():
            store.start_background_load()
        return JsonResponse(
            {
                'available': False,
                'loading': True,
                'reason': 'Screen data is still loading. Please wait until preparation finishes, then try again.',
            },
            status=202,
        )

    key = cache_key(gene, guide_positions)
    figure_json = get_cached_figure(key)
    if figure_json is None:
        build_err = get_build_error(key)
        if build_err:
            return JsonResponse(
                {'available': False, 'reason': f'Could not build screen plot: {build_err}'},
                status=500,
            )
        started = _ensure_full_figure_build_started(gene, guide_positions, store)
        return JsonResponse(
            {
                'available': False,
                'building': is_building(key) or started,
                'reason': 'Screen plots are still being prepared. Please wait a moment and try again.',
            },
            status=202,
        )

    return JsonResponse({
        'available': True,
        'figure': figure_json,
        'gene': gene,
        'markers': list(SCREEN_MARKERS),
    })


def tutorial(request):
    return render(request, 'designer/tutorial.html')

def about(request):
    return render(request, 'designer/about.html')
